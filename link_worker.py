from resolver import resolve_dep

try:
    from xlrd import open_workbook
except (ImportError, ModuleNotFoundError):
    resolve_dep("xlrd")
    from xlrd import open_workbook

try:
    from xlutils.copy import copy
except (ImportError, ModuleNotFoundError):
    resolve_dep("xlutils")
    from xlutils.copy import copy

try:
    import xmltodict
except (ImportError, ModuleNotFoundError):
    resolve_dep("xmltodict")
    import xmltodict

try:
    from openpyxl.workbook import Workbook
except (ImportError, ModuleNotFoundError):
    resolve_dep("openpyxl")
    from openpyxl.workbook import Workbook


import subprocess
import os
from shutil import rmtree

def cvt_xls_to_xlsx(src_file_path, dst_file_path):
    book_xls = open_workbook(src_file_path)
    book_xlsx = Workbook()
    sheet_names = book_xls.sheet_names()
    for sheet_index, sheet_name in enumerate(sheet_names):
        sheet_xls = book_xls.sheet_by_name(sheet_name)
        if sheet_index == 0:
            sheet_xlsx = book_xlsx.active
            sheet_xlsx.title = sheet_name
        else:
            sheet_xlsx = book_xlsx.create_sheet(title=sheet_name)

        for row in range(0, sheet_xls.nrows):
            for col in range(0, sheet_xls.ncols):
                sheet_xlsx.cell(row = row+1 , column = col+1).value = sheet_xls.cell_value(row, col)

    book_xlsx.save(dst_file_path)

def get_links(file_name):

    cvt_xls_to_xlsx(file_name, file_name[:-3] + 'xlsx')

    dir_path = os.path.dirname(os.path.realpath(file_name))

    subprocess.call(["unzip",str(file_name+"x"),"-d","file_xml"], stdout=open(os.devnull, 'wb'))


    xml_sheet_names = dict()

    with open_workbook(file_name,formatting_info=True) as rb:
        copy(rb)
        workbook_names_list = rb.sheet_names()
        for i,name in enumerate(workbook_names_list):
            xml_sheet_names[name] = "sheet"+str(i+1)

    links = dict()
    for i, k in enumerate(workbook_names_list):
        xmlFile = os.path.join(dir_path,"file_xml/xl/worksheets/{}.xml".format(xml_sheet_names[k]))
        with open(xmlFile) as f:
            xml = f.read()
        #print(json.dumps(xmltodict.parse(xml)['worksheet']["sheetData"]["row"], indent=4))
        n = 0
        for row in xmltodict.parse(xml)['worksheet']["sheetData"]["row"]:
            n += 1
            if n == 1:
                continue
            try:
                links[n] = row["c"][-1]['f']['#text']
            except:
                continue
        rmtree("file_xml/")
        return links
